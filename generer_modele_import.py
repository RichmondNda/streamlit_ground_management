"""
Script pour générer un fichier Excel exemple pour l'import de cotisations
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Créer des données exemple
data = {
    'nom': ['Dupont', 'Martin', 'Bernard', 'Legrand', 'Dubois'],
    'prenom': ['Jean', 'Marie', 'Pierre', 'Sophie', 'Luc'],
    'nombre_terrains': [3, 2, 1, 4, 2],
    '2025-08': [3000, 2000, 1000, 4000, 2000],
    '2025-09': [3000, 2000, 1000, 4000, 2000],
    '2025-10': [3000, 2000, 1000, 4000, 2000],
    '2025-11': [3000, 2000, 1000, 4000, 2000],
    '2025-12': [3000, 2000, 1000, 4000, 2000],
    '2026-01': [3000, 2000, 1000, 4000, 2000],
}

# Créer le DataFrame
df = pd.DataFrame(data)

# Nom du fichier
filename = 'modele_import_cotisations.xlsx'

# Créer le fichier Excel
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Cotisations', index=False)
    
    # Récupérer le workbook et la feuille
    workbook = writer.book
    worksheet = writer.sheets['Cotisations']
    
    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    info_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Styliser l'en-tête
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Styliser les données
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colonnes de montants en jaune clair
            if cell.column > 3:
                cell.fill = info_fill
    
    # Ajuster la largeur des colonnes
    worksheet.column_dimensions['A'].width = 15  # nom
    worksheet.column_dimensions['B'].width = 15  # prenom
    worksheet.column_dimensions['C'].width = 18  # nombre_terrains
    
    for i in range(4, worksheet.max_column + 1):
        col_letter = worksheet.cell(row=1, column=i).column_letter
        worksheet.column_dimensions[col_letter].width = 12
    
    # Figer les volets
    worksheet.freeze_panes = 'D2'

# Créer une feuille d'instructions
workbook = load_workbook(filename)

# Créer une nouvelle feuille pour les instructions
ws_instructions = workbook.create_sheet('Instructions', 0)

instructions = [
    ['📋 INSTRUCTIONS D\'IMPORT', ''],
    ['', ''],
    ['Format du fichier :', ''],
    ['', ''],
    ['Colonne A', 'nom - Nom du participant (OBLIGATOIRE)'],
    ['Colonne B', 'prenom - Prénom du participant (OBLIGATOIRE)'],
    ['Colonne C', 'nombre_terrains - Nombre de terrains possédés (OPTIONNEL)'],
    ['Colonnes D+', 'Format YYYY-MM (ex: 2025-08 pour Août 2025)'],
    ['', 'Montant de la cotisation en FCFA'],
    ['', ''],
    ['⚠️ IMPORTANT :', ''],
    ['', ''],
    ['1.', 'Les colonnes nom et prenom sont OBLIGATOIRES'],
    ['2.', 'Le format des dates doit être YYYY-MM (ex: 2025-08, 2026-01)'],
    ['3.', 'Les montants doivent être >= 500 FCFA (minimum)'],
    ['4.', 'Laissez vide les mois sans cotisation'],
    ['5.', 'Si nombre_terrains > 1, le montant sera divisé par terrain'],
    ['', ''],
    ['💡 Exemple :', ''],
    ['', ''],
    ['', 'Jean Dupont a 3 terrains et paie 3000 FCFA en Janvier 2026'],
    ['', '→ 3 cotisations seront créées : 1000 FCFA par terrain'],
    ['', ''],
    ['📊 Données exemple :', ''],
    ['', ''],
    ['', 'La feuille "Cotisations" contient des données exemple'],
    ['', 'Vous pouvez les modifier ou supprimer et ajouter vos propres données'],
    ['', ''],
    ['🚀 Pour importer :', ''],
    ['', ''],
    ['1.', 'Remplissez la feuille "Cotisations" avec vos données'],
    ['2.', 'Enregistrez le fichier'],
    ['3.', 'Dans l\'application, allez dans "📥 Import Excel"'],
    ['4.', 'Téléversez ce fichier'],
    ['5.', 'Choisissez si les cotisations sont déjà payées ou non'],
    ['6.', 'Cliquez sur "Importer"'],
]

# Ajouter les instructions
for row_idx, (col1, col2) in enumerate(instructions, start=1):
    cell_a = ws_instructions.cell(row=row_idx, column=1, value=col1)
    cell_b = ws_instructions.cell(row=row_idx, column=2, value=col2)
    
    # Styliser les titres
    if 'INSTRUCTIONS' in col1 or 'IMPORTANT' in col1 or 'Exemple' in col1 or 'Données' in col1 or 'Pour importer' in col1:
        cell_a.font = Font(bold=True, size=14, color='2F5496')
    elif col1 in ['Format du fichier :', 'Colonne A', 'Colonne B', 'Colonne C', 'Colonnes D+']:
        cell_a.font = Font(bold=True)
    elif col1 in ['1.', '2.', '3.', '4.', '5.', '6.']:
        cell_a.font = Font(bold=True, color='C65911')
        cell_b.font = Font(color='C65911')

# Ajuster la largeur des colonnes
ws_instructions.column_dimensions['A'].width = 20
ws_instructions.column_dimensions['B'].width = 70

# Sauvegarder
workbook.save(filename)

print(f"✅ Fichier Excel créé avec succès : {filename}")
print(f"📄 Contenu :")
print(f"   - Feuille 1 : Instructions détaillées")
print(f"   - Feuille 2 : Cotisations (données exemple à modifier)")
print(f"")
print(f"💡 Format des colonnes :")
print(f"   nom | prenom | nombre_terrains | 2025-08 | 2025-09 | 2025-10 | ...")
print(f"")
print(f"📊 5 participants exemple inclus avec cotisations d'Août 2025 à Janvier 2026")
