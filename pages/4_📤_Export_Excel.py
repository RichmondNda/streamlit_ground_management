"""
Page Export Excel
"""

import streamlit as st
import sqlite3
import pandas as pd
import io
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import DB_NAME
from constants import MOIS_NOMS
from auth import require_authentication, show_logout_button

# Configuration de la page
st.set_page_config(
    page_title="Export Excel - MEDD",
    page_icon="📤",
    layout="wide"
)

# Vérifier l'authentification
require_authentication()

# Afficher le bouton de déconnexion
show_logout_button()

# ============================================================================
# REQUÊTES POUR L'EXPORT
# ============================================================================

def get_available_years():
    """Récupère la liste des années disponibles dans les cotisations"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT annee FROM cotisations ORDER BY annee DESC")
    years = [row[0] for row in cursor.fetchall()]
    conn.close()
    return years


def generate_cotisations_report(start_date=None, end_date=None, participant_ids=None):
    """
    Génère un rapport des cotisations pour une période donnée
    
    Args:
        start_date: datetime - Date de début (par défaut: août 2025)
        end_date: datetime - Date de fin (par défaut: aujourd'hui)
        participant_ids: list - Liste des IDs de participants (None = tous)
    """
    # Date de début par défaut : août 2025
    if start_date is None:
        start_date = datetime(2025, 8, 1)
    # Date de fin par défaut : actuelle
    if end_date is None:
        end_date = datetime.now()
    
    # Générer la liste des mois
    months = []
    temp_date = start_date
    while temp_date <= end_date:
        months.append((temp_date.year, temp_date.month))
        temp_date = temp_date + relativedelta(months=1)
    
    # Récupérer les participants
    conn = sqlite3.connect(DB_NAME)
    
    if participant_ids and len(participant_ids) > 0:
        # Filtrer par participants sélectionnés
        placeholders = ','.join('?' * len(participant_ids))
        query = f"SELECT id, nom, prenom, nombre_terrains FROM participants WHERE id IN ({placeholders}) ORDER BY nom, prenom"
        participants = pd.read_sql_query(query, conn, params=participant_ids)
    else:
        # Tous les participants
        participants = pd.read_sql_query(
            "SELECT id, nom, prenom, nombre_terrains FROM participants ORDER BY nom, prenom", 
            conn
        )
    
    if participants.empty:
        conn.close()
        return None
    
    # Créer le dataframe de base
    result = participants.copy()
    
    # Ajouter une colonne pour chaque mois
    for year, month in months:
        col_name = f"{MOIS_NOMS[month-1]} {year}"
        
        # Récupérer les cotisations pour ce mois (somme par participant si plusieurs terrains)
        query = """
            SELECT participant_id, SUM(montant) as montant 
            FROM cotisations 
            WHERE annee = ? AND mois = ? AND paye = 1
            GROUP BY participant_id
        """
        cotis = pd.read_sql_query(query, conn, params=(year, month))
        
        # Fusionner avec le résultat
        if not cotis.empty:
            cotis.columns = ['id', col_name]
            result = result.merge(cotis, on='id', how='left')
        else:
            result[col_name] = None
    
    # Calculer le total par participant
    month_cols = [f"{MOIS_NOMS[m-1]} {y}" for y, m in months]
    result['TOTAL PAYÉ'] = result[month_cols].sum(axis=1)
    
    # Réorganiser les colonnes
    final_cols = ['nom', 'prenom', 'nombre_terrains'] + month_cols + ['TOTAL PAYÉ']
    result = result[final_cols]
    
    conn.close()
    return result, months


def export_to_excel(df, months):
    """Exporte le dataframe vers Excel avec mise en forme"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Cotisations")
        ws = writer.book["Cotisations"]
        
        # Styles
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        paid_fill = PatternFill("solid", fgColor="C6E0B4")
        total_fill = PatternFill("solid", fgColor="FFD966")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Styliser l'en-tête
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Styliser les cellules de données
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colonnes de montants (à partir de la colonne 3)
                if col_idx > 2 and col_idx < len(df.columns) + 1:
                    if cell.value:
                        cell.fill = paid_fill
                        cell.number_format = '#,##0'
                
                # Colonne TOTAL en jaune
                if col_idx == len(df.columns) + 1:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    if cell.value:
                        cell.number_format = '#,##0'
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20  # nom
        ws.column_dimensions['B'].width = 20  # prenom
        
        for i in range(3, len(df.columns) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        # Figer les volets (figer nom et prénom)
        ws.freeze_panes = 'C2'
    
    output.seek(0)
    return output


def export_cotisations_to_excel_pivot(start_date=None, end_date=None, participant_ids=None, only_paid=True):
    """
    Exporte les cotisations au format pivot avec style Excel
    
    Args:
        start_date: datetime - Date de début
        end_date: datetime - Date de fin
        participant_ids: list - Liste des IDs de participants
        only_paid: bool - Exporter uniquement les cotisations payées
    """
    conn = sqlite3.connect(DB_NAME)

    query = """
        SELECT p.nom, p.prenom, p.nombre_terrains, c.annee, c.mois, c.montant, c.paye
        FROM cotisations c
        JOIN participants p ON p.id = c.participant_id
        WHERE 1=1
    """

    params = []
    
    # Filtrer par période
    if start_date:
        query += " AND (c.annee > ? OR (c.annee = ? AND c.mois >= ?))"
        params.extend([start_date.year - 1, start_date.year, start_date.month])
    
    if end_date:
        query += " AND (c.annee < ? OR (c.annee = ? AND c.mois <= ?))"
        params.extend([end_date.year + 1, end_date.year, end_date.month])
    
    # Filtrer par participants
    if participant_ids and len(participant_ids) > 0:
        placeholders = ','.join('?' * len(participant_ids))
        query += f" AND p.id IN ({placeholders})"
        params.extend(participant_ids)

    if only_paid:
        query += " AND c.paye = 1"

    query += " ORDER BY p.nom, p.prenom, c.annee, c.mois"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    if df.empty:
        return None

    # Créer la colonne pivot format "YYYY-MM"
    df['col'] = df['annee'].astype(str) + "-" + df['mois'].astype(str).str.zfill(2)

    # Créer le tableau pivot
    pivot = df.pivot_table(
            index=['nom', 'prenom', 'nombre_terrains'],
            columns='col',
            values='montant',
            aggfunc='sum'
    ).reset_index()

    # Trier les colonnes de date
    date_cols = sorted([c for c in pivot.columns if '-' in str(c)])
    fixed_cols = ['nom', 'prenom', 'nombre_terrains']
    pivot = pivot[fixed_cols + date_cols]

    # Calculer le total par participant
    pivot['TOTAL'] = pivot[date_cols].sum(axis=1)

    # Créer le fichier Excel avec style
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pivot.to_excel(writer, index=False, sheet_name="Cotisations")
            ws = writer.book["Cotisations"]
            
            # Styles
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(bold=True, color="FFFFFF")
            paid_fill = PatternFill("solid", fgColor="C6E0B4")
            total_fill = PatternFill("solid", fgColor="FFD966")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Styliser l'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Styliser les cellules de données
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Colorer les cellules avec montant (colonnes date)
                    if col_idx > 3 and col_idx < len(pivot.columns) + 1:  # Colonnes de mois
                        if cell.value:
                            cell.fill = paid_fill
                            cell.number_format = '#,##0.00 €'
                    
                    # Colonne TOTAL en jaune
                    if col_idx == len(pivot.columns) + 1:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
                        cell.number_format = '#,##0.00 €'
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 20  # nom
            ws.column_dimensions['B'].width = 20  # prenom
            ws.column_dimensions['C'].width = 15  # nombre_terrains
            
            for i in range(4, len(pivot.columns) + 2):
                ws.column_dimensions[get_column_letter(i)].width = 12
            
            # Figer les volets
            ws.freeze_panes = 'D2'

    output.seek(0)
    return output


st.title("📤 Export Excel")

st.subheader("📋 Rapport des cotisations avec filtres")

# Filtres
with st.expander("🔍 Filtres d'export", expanded=True):
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**📅 Période**")
        date_debut = st.date_input(
            "Date de début",
            value=datetime(2025, 8, 1),
            min_value=datetime(2020, 1, 1),
            max_value=datetime.now(),
            format="DD/MM/YYYY",
            key="export_date_debut"
        )
        
    with col2:
        st.write("**📅 Fin**")
        date_fin = st.date_input(
            "Date de fin",
            value=datetime.now(),
            min_value=datetime(2020, 1, 1),
            max_value=datetime(2100, 12, 31),
            format="DD/MM/YYYY",
            key="export_date_fin"
        )
    
    st.divider()
    
    # Sélection des participants
    st.write("**👥 Participants**")
    conn_temp = sqlite3.connect(DB_NAME)
    all_participants = pd.read_sql_query(
        "SELECT id, nom, prenom FROM participants ORDER BY nom, prenom",
        conn_temp
    )
    conn_temp.close()
    
    if not all_participants.empty:
        participants_dict = {f"{row['nom']} {row['prenom']}": row['id'] 
                           for _, row in all_participants.iterrows()}
        
        tous_participants = st.checkbox(
            "Tous les participants",
            value=True,
            key="export_tous_participants"
        )
        
        if not tous_participants:
            selected_participants = st.multiselect(
                "Sélectionner les participants",
                options=list(participants_dict.keys()),
                key="export_selected_participants"
            )
            participant_ids_filter = [participants_dict[name] for name in selected_participants] if selected_participants else None
        else:
            participant_ids_filter = None
            selected_participants = None
    else:
        st.warning("Aucun participant dans la base de données")
        participant_ids_filter = None
        selected_participants = None

# Convertir les dates en datetime
start_datetime = datetime.combine(date_debut, datetime.min.time())
end_datetime = datetime.combine(date_fin, datetime.min.time())

# Validation
if start_datetime > end_datetime:
    st.error("⚠️ La date de début doit être antérieure à la date de fin")
    st.stop()

# Afficher les filtres actifs
st.info(f"📊 Export de **{start_datetime.strftime('%B %Y')}** à **{end_datetime.strftime('%B %Y')}** • "
        f"Participants : **{'Tous' if participant_ids_filter is None else f'{len(selected_participants)} sélectionné(s)'}**")

st.divider()

# Générer le rapport
with st.spinner("Génération du rapport..."):
    result = generate_cotisations_report(start_datetime, end_datetime, participant_ids_filter)

if result is not None:
    df, months = result
    
    # Afficher un aperçu du tableau
    st.write(f"**Aperçu du rapport** ({len(df)} participants, {len(months)} mois)")
    
    # Formater l'affichage pour remplacer les valeurs numériques
    df_display = df.copy()
    month_cols = [f"{MOIS_NOMS[m-1]} {y}" for y, m in months]
    
    for col in month_cols + ['TOTAL PAYÉ']:
        df_display[col] = df_display[col].apply(
            lambda x: f"{x:,.0f}".replace(',', ' ') + " FCFA" if pd.notna(x) and x > 0 else "-"
        )
    
    st.dataframe(df_display, use_container_width=True, height=400)
    
    # Statistiques rapides
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Participants", len(df))
    with col2:
        total_general = df['TOTAL PAYÉ'].sum()
        st.metric("Total général", f"{total_general:,.0f}".replace(',', ' ') + " FCFA")
    with col3:
        st.metric("Période", f"Août 2025 - {MOIS_NOMS[datetime.now().month-1]} {datetime.now().year}")
    
    st.divider()
    
    # Bouton d'export
    if st.button("📥 Générer et télécharger le fichier Excel", type="primary", use_container_width=True):
        with st.spinner("Génération du fichier Excel..."):
            excel = export_to_excel(df, months)
        
        st.success("✅ Fichier Excel généré avec succès")
        
        filename = f"rapport_cotisations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        st.download_button(
            label="💾 Télécharger le fichier Excel",
            data=excel,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
else:
    st.info("Aucun participant enregistré. Veuillez d'abord ajouter des participants.")

st.divider()

# Ancien export (conservé en option)
with st.expander("🔧 Export personnalisé (ancien format)"):
    st.subheader("Exporter les cotisations au format pivot")

st.divider()

# Export pivot avec les mêmes filtres
with st.expander("🔧 Export format pivot (optionnel)"):
    st.info("📌 Cet export utilise les mêmes filtres de période et de participants que ci-dessus")
    
    only_paid_pivot = st.checkbox("Uniquement les cotisations payées", value=True, key="only_paid_pivot")

    if st.button("Générer l'export pivot", type="secondary"):
        with st.spinner("Génération du fichier Excel pivot..."):
            excel = export_cotisations_to_excel_pivot(start_datetime, end_datetime, participant_ids_filter, only_paid_pivot)
        
        if excel:
            st.success("✅ Fichier Excel pivot généré avec succès")
            
            filename = f"cotisations_pivot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            st.download_button(
                label="📥 Télécharger le fichier Excel pivot",
                data=excel,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Aucune donnée à exporter avec ces filtres")
